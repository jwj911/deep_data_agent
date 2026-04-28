from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.merge import MergedCell
from typing import Any
import pandas as pd


def index_to_col_letter(n: int) -> str:
    """将数字索引转换为Excel列字母 (0 → 'A', 26 → 'AA')"""
    result = []
    while True:
        n, r = divmod(n, 26)
        result.append(chr(65 + r))  # 65 是 'A' 的 ASCII 码
        if n == 0:
            break
        n -= 1  # 调整进位，因为列标是 1-based 的 26 进制
    return "".join(reversed(result))


# 行列都从0开始
def find_word_pos(word_name, df):
    for row_idx in range(len(df)):
        for col_idx in range(len(df.columns)):
            cell_value = df.iloc[row_idx, col_idx]
            cell_value = str(cell_value).strip().replace("\n", "").replace(" ", "")
            if cell_value and str(cell_value) == word_name:
                return (row_idx, col_idx)

    return (None, None)


# 行列都从0开始
def set_cell_value(
    row: int = 0, col: int = 0, value: Any = None, sheet: Worksheet = None
):
    if isinstance(col, str):
        col = column_index_from_string(col)
    cell = sheet.cell(row=row, column=col)

    # 检查是否是合并单元格
    # if isinstance(cell, MergedCell):
    #     # 对于合并单元格，需要找到合并区域的左上角单元格
    #     for merged_range in sheet.merged_cells.ranges:
    #         if cell.coordinate in merged_range:
    #             # 获取合并区域的左上角单元格
    #             top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
    #             if top_left_cell.data_type == "f":
    #                 return
    #             top_left_cell.value = value
    #             return

    if cell.data_type == "f":
        return
    cell.value = value


def get_cell_value(row: int = 0, col: int = 0, sheet: Worksheet = None):
    if isinstance(col, str):
        col = column_index_from_string(col)
    return sheet.cell(row=row, column=col).value

def find_header_row(columns: list = [], df: pd.DataFrame = None):
    """查找包含指定列名的表头行"""
    # 准备不区分大小写的关键字
    columns_lower = [kw.strip().lower() for kw in columns]
    for index, row in df.iterrows():
        row_str = [
            (
                str(cell).lower().strip().replace("\n", "").replace(" ", "")
                if pd.notnull(cell)
                else ""
            )
            for cell in row
        ]
        if all(any(kw in cell for cell in row_str) for kw in columns_lower):
            return index
    return None

def get_df_by_columns(columns: list = [], df: pd.DataFrame = None):
    header_row = find_header_row(columns, df)
    if header_row is None:
        return None
    new_df = df.iloc[header_row + 1 :].copy()  # 从表头行的下一行开始取数据
    new_df.columns = df.iloc[header_row].values  # 将表头行设置为列名
    new_df.columns = [
        str(col).strip().replace("\n", "").replace(" ", "") for col in new_df.columns
    ]
    new_df.reset_index(drop=True, inplace=True)
    new_df.to_csv("middle_result.csv", index=False, encoding="utf-8-sig")
    return header_row, new_df


if __name__ == "__main__":
    letter = index_to_col_letter(n=0)
    print(letter)
